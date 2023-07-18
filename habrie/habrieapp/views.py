from datetime import date
from django.shortcuts import render
from rest_framework.response import Response
from django.utils.crypto import get_random_string
from rest_framework.decorators import api_view
from datetime import datetime
from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import MultiPartParser, JSONParser
from .models import Student, AcademicDetail
from .serializers import StudentSerializer, AcademicDetailSerializer, ParentSerializer
from .forms import *
from .tasks import *
from io import BytesIO
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from openpyxl import Workbook




def generate_enrollment_number(student_name):
    current_date = date.today().strftime('%d%m%y')
    initials = student_name[:3].upper()
    random_number = get_random_string(length=3, allowed_chars='0123456789')
    enrollment_number = current_date + initials + random_number
    return enrollment_number


@api_view(['POST'])
@parser_classes([MultiPartParser,JSONParser])
def create_student(request):
    upload_form = CsvForm(request.POST or None, request.FILES or None)

    if request.method == 'POST':

        data = request.data
        student_name = data['name']
        enrollment_number = generate_enrollment_number(student_name)
        academic_data = {
                'enroll_id': enrollment_number,
                'doj': datetime.strptime(request.data.get('doj'), '%Y-%m-%d').date(),
                'class_id': request.data['class_id'],
                'session': request.data['session'],
                'section_id': request.data['section_id'],
        }
        academic_serializer = AcademicDetailSerializer(data=academic_data)
        
        if not academic_serializer.is_valid():
            return Response({"message": "Academic Data not verified"}, status=400)
        
        if academic_serializer.is_valid():
            academic = academic_serializer.save()

        student_data = {
                'name': request.data.get('name'),
                'gender': request.data.get('gender'),
                'aadhar': request.data.get('aadhar'),
                'dob': datetime.strptime(request.data.get('dob'), '%Y-%m-%d').date(),
                'id_mark': request.data.get('id_mark'),
                'addmission_cat': request.data.get('addmission_cat'),
                'height': request.data.get('height'),
                'weight': request.data.get('weight'),
                'mail': request.data.get('mail'),
                'contact': request.data.get('contact'),
                'address': request.data.get('address'),
                'enroll_number' : academic.id
            }
        student_serializer = StudentSerializer(data=student_data)
        
        if not student_serializer.is_valid():
            return Response({"message": "Student Data not verified"}, status=400)


        if student_serializer.is_valid():
            student = student_serializer.save()

        parent_data = {
            'student_name' : student.id,
            'father_name' : request.data.get('father_name'),
            'father_qualification' : request.data.get('father_qualification'),
            'father_profession' : request.data.get('father_profession'),
            'father_designation' : request.data.get('father_designation'),
            'father_aadhar' : request.data.get('father_aadhar'),
            'father_number' : request.data.get('father_number'),
            'father_mail' : request.data.get('father_mail'),
            'mother_name' : request.data.get('father_mail'),
            'mother_qualification' : request.data.get('mother_qualification'),
            'mother_profession' : request.data.get('mother_profession'),
            'mother_designation' : request.data.get('mother_designation'),
            'mother_aadhar' : request.data.get('mother_aadhar'),
            'mother_number' : request.data.get('mother_number'),
            'mother_mail' : request.data.get('mother_mail'),
        }

        parent_serializer = ParentSerializer(data=parent_data)

        if not parent_serializer.is_valid():
            return Response({"message": "Parent Data not verified"}, status=400)

        if parent_serializer.is_valid():
            parent = parent_serializer.save()


        recipient_email = student_serializer.validated_data.get('mail')
        recipient_name = student_serializer.validated_data.get('name')
        enroll_no = enrollment_number
        class_id = academic_serializer.validated_data.get('class_id')
        section = academic_serializer.validated_data.get('section_id')
        session = academic_serializer.validated_data.get('session')

        send_templated_email.delay(recipient_email, recipient_name, enroll_no, class_id, section, session)

        if not upload_form.is_valid():
            return Response({"message": "Bulk data not verified"}, status=400)

        if upload_form.is_valid():
            print("YESS CSV")
            upload_form.save()
            csv_upload.delay()

        return Response({"message": "Single and bulk uploads done successfully"}, status=200)

    else:
        student_serializer = StudentSerializer()
        context = {
            'student_serializer': student_serializer,
            'upload_form': upload_form,
        }

    html_content = render(request, 'index1.html', context).content
    return Response(html_content)





@api_view(['GET'])
def export_to_pdf(request):
    class_filter = request.GET.get('class_filter')
    section_filter = request.GET.get('section_filter')
    admission_cat_filter = request.GET.get('admission_cat_filter')

    students = Student.objects.all()
    if class_filter:
        students = students.filter(enroll_number__class_id=class_filter)
    if section_filter:
        students = students.filter(enroll_number__section_id=section_filter)
    if admission_cat_filter:
        students = students.filter(addmission_cat=admission_cat_filter)

    template = get_template('pdf_template.html')
    context = {
        'students': students,
        'class_filter': class_filter,
        'section_filter': section_filter,
        'admission_cat_filter': admission_cat_filter,
    }
    html = template.render(context)

    pdf_file = BytesIO()
    pisa.CreatePDF(html, dest=pdf_file)

    response = HttpResponse(pdf_file.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="student_list.pdf"'

    return response





@api_view(['GET'])
def export_to_excel(request):
    class_filter = request.GET.get('class_filter')
    section_filter = request.GET.get('section_filter')
    admission_cat_filter = request.GET.get('admission_cat_filter')

    students = Student.objects.all()
    if class_filter:
        students = students.filter(enroll_number__class_id=class_filter)
    if section_filter:
        students = students.filter(enroll_number__section_id=section_filter)
    if admission_cat_filter:
        students = students.filter(addmission_cat=admission_cat_filter)

    workbook = Workbook()
    sheet = workbook.active

    column_names = ['Name', 'Gender', 'Aadhar', 'Date of Birth', 'Admission Category', 'Class', 'Section', 'Height', 'Weight', 'Mail', 'Contact', 'Address']

    for index, column_name in enumerate(column_names, start=1):
        sheet.cell(row=1, column=index, value=column_name)

    for row, student in enumerate(students, start=2):
        sheet.cell(row=row, column=1, value=student.name)
        sheet.cell(row=row, column=2, value=student.gender)
        sheet.cell(row=row, column=3, value=student.aadhar)
        sheet.cell(row=row, column=4, value=student.dob)
        sheet.cell(row=row, column=5, value=student.addmission_cat)
        sheet.cell(row=row, column=6, value=student.enroll_number.class_id)
        sheet.cell(row=row, column=7, value=student.enroll_number.section_id)
        sheet.cell(row=row, column=8, value=student.height)
        sheet.cell(row=row, column=9, value=student.weight)
        sheet.cell(row=row, column=10, value=student.mail)
        sheet.cell(row=row, column=11, value=student.contact)
        sheet.cell(row=row, column=12, value=student.address)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="filtered_student_list.xlsx"'

    workbook.save(response)

    return response


@api_view(['POST'])
def upload_document(request):
    if request.method == 'POST':
        form = DocumentForm(request.data, request.FILES)
        if form.is_valid():
            form.save()
            return Response({"message": "Document uploaded successfully"}, status=201)
        else:
            return Response({"error": "Invalid form data", "details": form.errors}, status=400)
    else:
        form = DocumentForm()
    
    context = {'form': form}
    return Response(context, template_name='document_upload.html')
